#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MeterSphere 用例格式转换脚本
将多步骤用例（每个步骤一行）转换为合并格式（所有步骤合并在一个单元格）

功能:
    - 将同一用例的多个步骤行合并为一个单元格
    - 自动去除原有编号（如"1、"、"2."、"3）"等），使用统一的"[编号]"格式
    - 动态适配源文件字段，只输出源文件中存在的字段
    - 支持跨平台运行（Windows/Linux）

用法:
    python convert_metersphere_case.py <输入文件> [输出文件]

示例:
    python convert_metersphere_case.py Metersphere_case_XT邮件系统.xlsx
    python convert_metersphere_case.py Metersphere_case_XT邮件系统.xlsx output.xlsx

依赖:
    pip install openpyxl
"""

import sys
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment


# MeterSphere支持的所有字段（按推荐输出顺序）
ALL_FIELDS = [
    "ID",
    "用例名称",
    "所属模块",
    "前置条件",
    "备注",
    "步骤描述",
    "预期结果",
    "编辑模式",
    "标签",
    "用例状态",
    "责任人",
    "用例等级",
    "版本",
    "验收用例",
    "测试方式",
    "自动化实现",
    "自动化类名",
    "JIRA号",
    "用例类型",
    "自动化说明",
    "评论",
    "执行结果",
    "评审结果",
    "创建人",
    "创建时间",
    "更新人",
    "更新时间",
]

# 必需字段（用于判断用例边界和步骤合并）
REQUIRED_FIELDS = ["ID", "步骤描述", "预期结果"]


def remove_existing_numbering(text: str) -> str:
    """去除文本开头的已有编号格式"""
    if not text:
        return text
    pattern = r"^\s*\d+\s*[、.．）)]\s*"
    return re.sub(pattern, "", text)


def get_source_headers(sheet) -> list:
    """读取源文件表头"""
    headers = []
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value:
            headers.append(value)
        else:
            headers.append(None)
    return headers


def build_column_map(headers: list) -> dict:
    """构建字段名到列索引的映射"""
    col_map = {}
    for idx, header in enumerate(headers, 1):
        if header and header in ALL_FIELDS:
            col_map[header] = idx
    return col_map


def determine_output_fields(source_fields: list) -> list:
    """确定输出字段：只包含源文件中存在的字段，按推荐顺序排列"""
    output = []
    for field in ALL_FIELDS:
        if field in source_fields:
            output.append(field)
    return output


def convert_case_format(input_file: str, output_file: str = None):
    """转换用例格式"""
    if output_file is None:
        base_name = os.path.splitext(input_file)[0]
        output_file = f"{base_name}_converted.xlsx"

    print(f"正在读取源文件: {input_file}")
    wb_source = load_workbook(input_file)
    sheet_source = wb_source.active

    # 读取源文件表头
    source_headers = get_source_headers(sheet_source)
    source_fields = [h for h in source_headers if h in ALL_FIELDS]
    col_map = build_column_map(source_headers)

    print(f"检测到字段: {source_fields}")

    # 检查必需字段
    missing = [f for f in REQUIRED_FIELDS if f not in col_map]
    if missing:
        print(f"错误: 缺少必需字段: {missing}")
        return None

    # 确定输出字段
    output_fields = determine_output_fields(source_fields)
    print(f"输出字段: {output_fields}")

    # 创建新工作簿
    wb_target = Workbook()
    sheet_target = wb_target.active
    sheet_target.title = "模版"

    # 写入表头
    for col_idx, header in enumerate(output_fields, 1):
        cell = sheet_target.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 解析源数据，按用例分组
    cases = []
    current_case = None

    for row_idx in range(2, sheet_source.max_row + 1):
        case_id_col = col_map["ID"]
        case_id = sheet_source.cell(row=row_idx, column=case_id_col).value

        if case_id is not None:
            # 新用例开始
            if current_case is not None:
                cases.append(current_case)

            current_case = {"_steps": [], "_expected": []}
            # 读取所有字段值
            for field in source_fields:
                if field in col_map:
                    current_case[field] = sheet_source.cell(
                        row=row_idx, column=col_map[field]
                    ).value

        # 添加步骤和预期结果
        if current_case is not None:
            step = sheet_source.cell(row=row_idx, column=col_map["步骤描述"]).value
            expected = sheet_source.cell(row=row_idx, column=col_map["预期结果"]).value

            if step:
                current_case["_steps"].append(step)
            if expected:
                current_case["_expected"].append(expected)

    # 添加最后一个用例
    if current_case is not None:
        cases.append(current_case)

    print(f"共解析到 {len(cases)} 条用例")

    # 写入转换后的数据
    for case_idx, case in enumerate(cases, 2):
        # 合并步骤和预期结果
        merged_steps = "\n".join(
            f"[{i}]{remove_existing_numbering(step)}"
            for i, step in enumerate(case["_steps"], 1)
        )
        merged_expected = "\n".join(
            f"[{i}]{remove_existing_numbering(result)}"
            for i, result in enumerate(case["_expected"], 1)
        )

        # 写入行数据
        for col_idx, field in enumerate(output_fields, 1):
            if field == "步骤描述":
                value = merged_steps
            elif field == "预期结果":
                value = merged_expected
            elif field == "编辑模式":
                value = "STEP"
            else:
                value = case.get(field)

            cell = sheet_target.cell(row=case_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 调整列宽
    column_width_map = {
        "ID": 12,
        "用例名称": 40,
        "所属模块": 25,
        "前置条件": 30,
        "备注": 15,
        "步骤描述": 50,
        "预期结果": 50,
        "编辑模式": 10,
        "标签": 10,
        "用例状态": 10,
        "责任人": 10,
        "用例等级": 10,
        "版本": 10,
        "验收用例": 10,
        "测试方式": 10,
        "自动化实现": 15,
        "自动化类名": 15,
        "JIRA号": 15,
        "用例类型": 10,
        "自动化说明": 15,
        "评论": 20,
        "执行结果": 10,
        "评审结果": 10,
        "创建人": 10,
        "创建时间": 15,
        "更新人": 10,
        "更新时间": 15,
    }

    for col_idx, field in enumerate(output_fields, 1):
        col_letter = chr(64 + col_idx)  # A=65, 所以64+1=A
        width = column_width_map.get(field, 15)
        sheet_target.column_dimensions[col_letter].width = width

    # 设置行高
    for row_idx in range(2, len(cases) + 2):
        sheet_target.row_dimensions[row_idx].height = 80

    # 保存文件
    wb_target.save(output_file)
    print(f"转换完成，输出文件: {output_file}")

    return output_file


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("错误: 请指定输入文件")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(input_file):
        print(f"错误: 输入文件不存在: {input_file}")
        sys.exit(1)

    convert_case_format(input_file, output_file)


if __name__ == "__main__":
    main()
